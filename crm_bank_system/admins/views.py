from decimal import Decimal

from django.db import transaction, connection
from django.db.models import Count, Sum

from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.admin.views.decorators import staff_member_required
from django.views.generic import TemplateView

import openpyxl
from dulwich.porcelain import status

from transactions.models import Transaction

from users.models import CustomUser

from .forms import AdminUserEditForm, RespondToReportForm
from trades.models import Trades

from documents.models import Report

from django.contrib import messages

from django.urls import reverse

from django.db.models import F

import json
from django.http import HttpResponse

from customers.models import ClientCard

from users.models import UserClientCard

from .services import *

# Create your views here.

class AdminTransactionListView(UserPassesTestMixin, TemplateView):
    template_name = "admin/transactions_list.html"

    def test_func(self):
        return self.request.user.is_staff

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        request = self.request

        search_query = request.GET.get("search", "").strip()
        type_filter = request.GET.get("type", "").strip()

        transactions = Transaction.objects.all().order_by("-created_at")

        if search_query:
            transactions = transactions.filter(sender__username__icontains=search_query)

        if type_filter:
            transactions = transactions.filter(type=type_filter)

        context["transactions"] = transactions
        context["type_choices"] = Transaction.TYPE_CHOICES
        return context


class AdminUserListView(UserPassesTestMixin, View):
    template_name = "admin/users_list.html"

    def test_func(self):
        return self.request.user.is_staff

    def get(self, request, *args, **kwargs):
        search_query = request.GET.get("search", "").strip()

        query = """
            SELECT id, username, email, first_name, last_name, date_joined, last_login, salary, trades_count, is_active,  is_superuser
            FROM "Users"
        """

        params = []

        if search_query:
            query += """
                WHERE username ILIKE %s
            """
            search_query = f"%{search_query}%"
            params.append(search_query)

        query += """
            ORDER BY username
        """

        # Execute the query
        with connection.cursor() as cursor:
            cursor.execute(query, params)
            users_data = cursor.fetchall()

        users = []
        for row in users_data:
            users.append({
                'id': row[0],
                'username': row[1],
                'email': row[2],
                'first_name': row[3],
                'last_name': row[4],
                'date_joined': row[5],
                'last_login': row[6],
                'salary': row[7],
                'trades_count': row[8],
                'is_active': row[10],
                'is_superuser': row[9],
            })

        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return render(request, "admin/users_list.html", {"users": users})

        return render(request, self.template_name, {"users": users})


class ExportUsersToExcelView(View):
    def get(self, request, *args, **kwargs):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Users"

        headers = ["ID", "Username", "Email", "Date Joined", "Last Login", "Зарплата", "Счетчик трейдов", "Админ", "Имеет доступ"]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Fetch all users (main query)
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, username, email, date_joined, last_login, salary, trades_count, is_superuser, is_active
                FROM "Users"
            """)
            users_data = cursor.fetchall()

        # Populate main users data
        for row_num, user in enumerate(users_data, start=2):
            sheet.cell(row=row_num, column=1, value=user[0])  # ID
            sheet.cell(row=row_num, column=2, value=user[1])  # Username
            sheet.cell(row=row_num, column=3, value=user[2])  # Email
            sheet.cell(row=row_num, column=4, value=user[3].strftime('%Y-%m-%d %H:%M:%S') if user[3] else "")  # Date Joined
            sheet.cell(row=row_num, column=5, value=user[4].strftime('%Y-%m-%d %H:%M:%S') if user[4] else "")  # Last Login
            sheet.cell(row=row_num, column=6, value=user[5])  # Зарплата
            sheet.cell(row=row_num, column=7, value=user[6])  # Счетчик трейдов
            sheet.cell(row=row_num, column=8, value="True" if user[7] else "False")  # Админ
            sheet.cell(row=row_num, column=9, value="True" if user[8] else "False")  # Имеет доступ

        # Fetch users sorted by trade count (descending)
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, username, email, date_joined, last_login, salary, trades_count, is_superuser, is_active
                FROM "Users"
                ORDER BY trades_count DESC
            """)
            users_by_trades = cursor.fetchall()

        # Write users sorted by trade count
        trades_start_row = len(users_data) + 3
        sheet.cell(row=trades_start_row, column=1, value="Пользователи, отсортированные по счетчику трейдов (по убыванию)")
        for row_num, user in enumerate(users_by_trades, start=trades_start_row + 1):
            sheet.cell(row=row_num, column=1, value=user[0])  # ID
            sheet.cell(row=row_num, column=2, value=user[1])  # Username
            sheet.cell(row=row_num, column=3, value=user[2])  # Email
            sheet.cell(row=row_num, column=4, value=user[3].strftime('%Y-%m-%d %H:%M:%S') if user[3] else "")  # Date Joined
            sheet.cell(row=row_num, column=5, value=user[4].strftime('%Y-%m-%d %H:%M:%S') if user[4] else "")  # Last Login
            sheet.cell(row=row_num, column=6, value=user[5])  # Зарплата
            sheet.cell(row=row_num, column=7, value=user[6])  # Счетчик трейдов
            sheet.cell(row=row_num, column=8, value="True" if user[7] else "False")  # Админ
            sheet.cell(row=row_num, column=9, value="True" if user[8] else "False")  # Имеет доступ

        # Fetch admins
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, username, email, date_joined, last_login, salary, trades_count, is_superuser, is_active
                FROM "Users"
                WHERE is_superuser = TRUE
            """)
            admins_data = cursor.fetchall()

        # Write admins
        admins_start_row = trades_start_row + len(users_by_trades) + 2
        sheet.cell(row=admins_start_row, column=1, value="Админы")
        for row_num, user in enumerate(admins_data, start=admins_start_row + 1):
            sheet.cell(row=row_num, column=1, value=user[0])  # ID
            sheet.cell(row=row_num, column=2, value=user[1])  # Username
            sheet.cell(row=row_num, column=3, value=user[2])  # Email
            sheet.cell(row=row_num, column=4, value=user[3].strftime('%Y-%m-%d %H:%M:%S') if user[3] else "")  # Date Joined
            sheet.cell(row=row_num, column=5, value=user[4].strftime('%Y-%m-%d %H:%M:%S') if user[4] else "")  # Last Login
            sheet.cell(row=row_num, column=6, value=user[5])  # Зарплата
            sheet.cell(row=row_num, column=7, value=user[6])  # Счетчик трейдов
            sheet.cell(row=row_num, column=8, value="True" if user[7] else "False")  # Админ
            sheet.cell(row=row_num, column=9, value="True" if user[8] else "False")  # Имеет доступ

        # Fetch users sorted by salary (descending)
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, username, email, date_joined, last_login, salary, trades_count, is_superuser, is_active
                FROM "Users"
                ORDER BY salary DESC
            """)
            users_by_salary = cursor.fetchall()

        # Write users sorted by salary
        salary_start_row = admins_start_row + len(admins_data) + 2
        sheet.cell(row=salary_start_row, column=1, value="Пользователи, отсортированные по зарплате (по убыванию)")
        for row_num, user in enumerate(users_by_salary, start=salary_start_row + 1):
            sheet.cell(row=row_num, column=1, value=user[0])  # ID
            sheet.cell(row=row_num, column=2, value=user[1])  # Username
            sheet.cell(row=row_num, column=3, value=user[2])  # Email
            sheet.cell(row=row_num, column=4, value=user[3].strftime('%Y-%m-%d %H:%M:%S') if user[3] else "")  # Date Joined
            sheet.cell(row=row_num, column=5, value=user[4].strftime('%Y-%m-%d %H:%M:%S') if user[4] else "")  # Last Login
            sheet.cell(row=row_num, column=6, value=user[5])  # Зарплата
            sheet.cell(row=row_num, column=7, value=user[6])  # Счетчик трейдов
            sheet.cell(row=row_num, column=8, value="True" if user[7] else "False")  # Админ
            sheet.cell(row=row_num, column=9, value="True" if user[8] else "False")  # Имеет доступ

        # Prepare response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="users.xlsx"'
        workbook.save(response)

        return response

class UserEditView(UserPassesTestMixin, View):
    template_name = "admin/user_edit.html"

    def test_func(self):
        return self.request.user.is_staff

    def get(self, request, pk):
        user = CustomUser.objects.get(id=pk)
        form = AdminUserEditForm(instance=user)
        return render(request, self.template_name, {"form": form, "user": user})

    def post(self, request, pk):
        user = CustomUser.objects.get(id=pk)
        form = AdminUserEditForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            return redirect('users_list')
        return render(request, self.template_name, {"form": form, "user": user})


class AdminTradesListView(UserPassesTestMixin, TemplateView):
    template_name = 'admin/admin_trades_list.html'

    def test_func(self):
        return self.request.user.is_staff

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        search_query = self.request.GET.get('search', '').strip()
        status_filter = self.request.GET.get('status', '').strip()

        trades = Trades.objects.all().order_by('-created_at')

        if search_query:
            trades = trades.filter(author__username__icontains=search_query)

        if status_filter:
            trades = trades.filter(status=status_filter)

        context['trades'] = trades
        context['status_choices'] = Trades.STATUS_CHOICES
        return context


class AdminFreezeTradeView(UserPassesTestMixin, View):
    def test_func(self):
        return self.request.user.is_staff

    def post(self, request, trade_id, *args, **kwargs):
        with transaction.atomic():
            trade = get_object_or_404(Trades, id=trade_id)

            if trade.status != "ACTIVE":
                messages.error(request, "Невозможно заморозить трейд с текущим статусом.")
                return redirect(reverse("admin_trades_list"))

            card = trade.card

            if card is None:
                messages.error(request, "Карта для трейда не найдена.")
                return redirect(reverse("admin_trades_list"))

            if trade.currency_to != "USD":
                current_balance = Decimal(card.balance.get(trade.currency_to, "0"))
                card.balance[trade.currency_to] = str(current_balance + trade.amount_to)
            else:
                card.balance_in_usd += trade.amount_to

            card.save()

            trade.status = "FROZEN"

            trade.author.trades_count = F("trades_count") - 1
            trade.author.save()

            trade.save()
            messages.success(request, f"Трейд #{trade.id} был заморожен.")

            return redirect(reverse("admin_trades_list"))


class AdminActivateTradeView(UserPassesTestMixin, View):
    def test_func(self):
        return self.request.user.is_staff

    def post(self, request, trade_id, *args, **kwargs):
        with transaction.atomic():
            trade = get_object_or_404(Trades, id=trade_id)

            if not trade.author.has_permanent_discount and trade.author.trades_count >= 5:
                raise ValueError(f"Количество трейдов обычного пользователя не может быть больше 5")

            else:
                if trade.status != "FROZEN":
                    messages.error(request, "Невозможно активировать трейд с текущим статусом.")
                    return redirect(reverse("admin_trades_list"))

                card = trade.card

                if card is None:
                    messages.error(request, "Карта для трейда не найдена.")
                    return redirect(reverse("admin_trades_list"))

                if trade.currency_to != "USD":
                    current_balance = Decimal(card.balance.get(trade.currency_to, "0"))  # Преобразуем строку в Decimal
                    if current_balance < trade.amount_to:
                        messages.error(request, "Недостаточно средств на балансе.")
                        return redirect(reverse("admin_trades_list"))

                    card.balance[trade.currency_to] = str(current_balance - trade.amount_to)  # Сохраняем как строку
                else:
                    card.balance_in_usd -= trade.amount_to

                card.save()

                trade.status = "ACTIVE"
                trade.author.trades_count = F("trades_count") + 1
                trade.author.save()

                trade.save()
                messages.success(request, f"Трейд #{trade.id} был активирован.")

                return redirect(reverse("admin_trades_list"))


class ExportTransactionsToExcelView(View):
    def get(self, request, *args, **kwargs):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Transactions"

        headers = ["ID", "Sender", "Recipient", "Amount", "Type", "Status", "Date"]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, sender_id, recipient_id, amount, type, status, created_at
                FROM "Transactions"
            """)
            transactions_data = cursor.fetchall()

        for row_num, transaction in enumerate(transactions_data, start=2):
            sheet.cell(row=row_num, column=1, value=transaction[0])
            sheet.cell(row=row_num, column=2, value=transaction[1])
            sheet.cell(row=row_num, column=3, value=transaction[2] if transaction[2] else "")
            sheet.cell(row=row_num, column=4, value=Decimal(str(transaction[3])))
            sheet.cell(row=row_num, column=5, value=transaction[4])
            sheet.cell(row=row_num, column=6, value=transaction[5])
            # Convert created_at to naive datetime
            created_at_naive = transaction[6].replace(tzinfo=None)
            sheet.cell(row=row_num, column=7, value=created_at_naive)

        with connection.cursor() as cursor:
            cursor.execute('SELECT COUNT(*) FROM "Transactions"')
            total_transactions = cursor.fetchone()[0]

        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT type, COUNT(*) as count
                FROM "Transactions"
                GROUP BY type
                ORDER BY count DESC
                LIMIT 1
            """)
            popular_transaction_type = cursor.fetchone()

        with connection.cursor() as cursor:
            cursor.execute('SELECT SUM(amount) AS total FROM "Transactions"')
            total_amount = cursor.fetchone()[0] or 0

        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT type, SUM(amount) AS total_amount
                FROM "Transactions"
                GROUP BY type
                ORDER BY type
            """)
            amounts_by_type_data = cursor.fetchall()

        summary_start_row = len(transactions_data) + 3
        sheet.cell(row=summary_start_row + 1, column=1, value="Общее количество транзакций")
        sheet.cell(row=summary_start_row + 1, column=2, value=total_transactions)

        if popular_transaction_type:
            sheet.cell(row=summary_start_row + 2, column=1, value="Наиболее популярный тип транзакций")
            sheet.cell(row=summary_start_row + 2, column=2, value=popular_transaction_type[0])

        sheet.cell(row=summary_start_row + 3, column=1, value="Общая сумма всех транзакций")
        sheet.cell(row=summary_start_row + 3, column=2, value=Decimal(str(total_amount)))

        sheet.cell(row=summary_start_row + 5, column=1, value="Сумма транзакций по типам")
        type_row = summary_start_row + 6
        sheet.cell(row=type_row, column=1, value="Тип")
        sheet.cell(row=type_row, column=2, value="Сумма")
        for entry in amounts_by_type_data:
            type_row += 1
            sheet.cell(row=type_row, column=1, value=entry[0])
            sheet.cell(row=type_row, column=2, value=Decimal(str(entry[1])))

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="transactions.xlsx"'
        workbook.save(response)

        return response


class ExportTradesToExcelView(View):
    def get(self, request, *args, **kwargs):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Trades"

        headers = ["ID", "Type", "Author", "Recipient", "Amount from", "Currency from", "Amount to", "Currency to", "Date", "Status"]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Fetch all trades
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, trades_type, author_id, other_user_id, amount_from, currency_from, amount_to, currency_to, created_at, status
                FROM "Trades"
            """)
            trades_data = cursor.fetchall()

        # Populate trades data
        for row_num, trade in enumerate(trades_data, start=2):
            sheet.cell(row=row_num, column=1, value=trade[0])
            sheet.cell(row=row_num, column=2, value=trade[1])
            sheet.cell(row=row_num, column=3, value=trade[2])
            sheet.cell(row=row_num, column=4, value=trade[3] if trade[3] else "")
            sheet.cell(row=row_num, column=5, value=Decimal(str(trade[4])))
            sheet.cell(row=row_num, column=6, value=trade[5])
            sheet.cell(row=row_num, column=7, value=Decimal(str(trade[6])))
            sheet.cell(row=row_num, column=8, value=trade[7])
            sheet.cell(row=row_num, column=10, value=trade[9])

            # Convert created_at to naive datetime
            created_at_naive = trade[8].replace(tzinfo=None)
            sheet.cell(row=row_num, column=9, value=created_at_naive)

        # Calculate total trades
        with connection.cursor() as cursor:
            cursor.execute('SELECT COUNT(*) FROM "Trades"')
            total_trades = cursor.fetchone()[0]

        # Find the most popular trade type
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT trades_type, COUNT(*) as count
                FROM "Trades"
                GROUP BY trades_type
                ORDER BY count DESC
                LIMIT 1
            """)
            popular_trade_type = cursor.fetchone()

        # Find the author who created the most trades
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT author_id, COUNT(*) as count
                FROM "Trades"
                GROUP BY author_id
                ORDER BY count DESC
                LIMIT 1
            """)
            top_author = cursor.fetchone()

        # Write summary data
        summary_start_row = len(trades_data) + 3
        sheet.cell(row=summary_start_row + 1, column=1, value="Общее количество трейдов")
        sheet.cell(row=summary_start_row + 1, column=2, value=total_trades)

        if popular_trade_type:
            sheet.cell(row=summary_start_row + 2, column=1, value="Наиболее популярный тип трейда")
            sheet.cell(row=summary_start_row + 2, column=2, value=popular_trade_type[0])

        if top_author:
            sheet.cell(row=summary_start_row + 3, column=1, value="Автор, создавший наибольшее количество трейдов")
            sheet.cell(row=summary_start_row + 3, column=2, value=top_author[0])

        # Prepare response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="trades.xlsx"'
        workbook.save(response)

        return response


class AdminPendingReportsView(UserPassesTestMixin, View):
    def test_func(self):
        return self.request.user.is_superuser

    def get(self, request):
        reports = Report.objects.filter(status='PENDING').order_by('-created_at')
        return render(request, 'documents/reports_list.html', {'reports': reports})


class RespondToReportView(UserPassesTestMixin, View):
    def test_func(self):
        return self.request.user.is_superuser

    def get(self, request, report_id):
        report = get_object_or_404(Report, id=report_id, status='PENDING')
        form = RespondToReportForm()
        return render(request, 'admin/respond_to_report.html', {'form': form, 'report': report})

    def post(self, request, report_id):
        report = get_object_or_404(Report, id=report_id, status='PENDING')
        form = RespondToReportForm(request.POST)
        if form.is_valid():
            report.status = 'RESOLVED'
            report.admin_response = form.cleaned_data['admin_response']
            report.save()
            return redirect('admin_pending_reports')
        return render(request, 'admin/respond_to_report.html', {'form': form, 'report': report})


class RejectReportView(UserPassesTestMixin, View):
    def test_func(self):
        return self.request.user.is_superuser

    def get(self, request, report_id):
        report = get_object_or_404(Report, id=report_id, status='PENDING')
        report.status = 'REJECTED'
        report.save()
        return redirect('admin_pending_reports')

class ExportReportJsonView(UserPassesTestMixin, View):
    def test_func(self):
        # Ensure the user is an admin
        return self.request.user.is_superuser

    def get(self, request, report_id):
        report = get_object_or_404(Report, id=report_id)

        report_data = {
            'id': report.id,
            'user': report.user.username,
            'description': report.description,
            'status': report.get_status_display(),
            'created_at': report.created_at.isoformat(),
            'updated_at': report.updated_at.isoformat(),
            'screenshot': report.screenshot,
            'document_url': report.document_url,
            'admin_response': report.admin_response,
        }

        json_data = json.dumps(report_data, ensure_ascii=False, indent=4)

        response = HttpResponse(json_data, content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="report_{report.id}.json"'

        return response


class ExportAllPendingReportsJsonView(UserPassesTestMixin, View):
    def test_func(self):
        # Ensure the user is an admin
        return self.request.user.is_superuser

    def get(self, request):
        # Fetch all pending reports
        pending_reports = Report.objects.filter(status='PENDING')

        # Serialize the reports data to JSON
        reports_data = []
        for report in pending_reports:
            report_data = {
                'id': report.id,
                'user': report.user.username,
                'description': report.description,
                'status': report.get_status_display(),
                'created_at': report.created_at.isoformat(),
                'updated_at': report.updated_at.isoformat(),
                'screenshot': report.screenshot,
                'document_url': report.document_url,
                'admin_response': report.admin_response,
            }
            reports_data.append(report_data)

        # Convert the data to JSON format
        json_data = json.dumps(reports_data, ensure_ascii=False, indent=4)

        # Create an HTTP response with the JSON data
        response = HttpResponse(json_data, content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename="pending_reports.json"'

        return response

class AdminClientCardsListView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        return self.request.user.is_superuser

    def get(self, request):
        sort = request.GET.get("sort", "asc")

        cards = ClientCard.objects.all()

        if sort == "asc":
            cards = cards.order_by("created_at")
        elif sort == "desc":
            cards = cards.order_by("-created_at")

        first_cards = {}
        for user in CustomUser.objects.all():
            first_card = ClientCard.objects.filter(user=user).order_by("created_at").first()
            if first_card:
                first_cards[user] = first_card

        for card in cards:
            card.is_first_card = card == first_cards.get(card.user)

        return render(request, "admin_client_cards_list.html", {
            "cards": cards,
            "sort": sort,
        })


class AdminDefaultCardsListView(LoginRequiredMixin, View, UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_superuser

    def get(self, request):
        sort = request.GET.get("sort", "asc")

        user_cards = UserClientCard.objects.all()

        if sort == "asc":
            user_cards = user_cards.order_by("created_at")
        elif sort == "desc":
            user_cards = user_cards.order_by("-created_at")

        first_cards = {
            user: UserClientCard.objects.filter(user=user).order_by("created_at").first()
            for user in CustomUser.objects.all()
        }

        for card in user_cards:
            card.is_first_card = card == first_cards.get(card.user)

        return render(request, "admin/admin_user_default_cards_list.html", {
            "cards": user_cards,
            "sort": sort,
        })


def crypto_growth_view(request):
    try:
        currencies = set_default_currencies_list()
        chart_data_list = {}

        for crypto in currencies:
            chart_data = process_data_for_single_currency(crypto)
            chart_data_list[crypto] = json.dumps(chart_data)

        return render(request, "crypto_growth.html", {"chart_data_list": chart_data_list})
    except Exception as e:
        return render(request, "crypto_growth.html", {"error": str(e)})


class AdminIndexView(LoginRequiredMixin, View, UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_superuser

    def get(self, request):
        return render(request, 'admin/admin_index.html')

